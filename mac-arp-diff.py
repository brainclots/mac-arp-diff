# $language = "python"
# $interface = "1.0"

'''
Purpose:    Collect the MAC address table on a switch and write it to an excel
            spreadsheet on one worksheet, then collect the ARP table and write
            it to a second worksheet. Copy in a worksheet of all IEEE OUIs
            and create formulae to display the manufacturer of each MAC, and
            highlight all the MAC addresses that do not have a corresponding
            IP address from the ARP table.
Author:
            ___  ____ _ ____ _  _    _  _ _    ____ ___ ___
            |__] |__/ | |__| |\ |    |_/  |    |  |  |    /
            |__] |  \ | |  | | \|    | \_ |___ |__|  |   /__
            Brian.Klotz@nike.com

Version:    1.0
Date:       June 2017
'''

import SecureCRT
import os
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import FormulaRule
import datetime

script_tab = crt.GetScriptTab()
script_tab.Screen.Synchronous = True
script_tab.Screen.IgnoreEscape = True


def main():
    screenRow = script_tab.Screen.CurrentRow
    screenCol = script_tab.Screen.CurrentColumn
    prompt = script_tab.Screen.Get(screenRow, 1, screenRow, screenCol).strip()
    if '(config' in prompt:
        crt.Dialog.MessageBox('Run script from user or priviliged exec only.')
        return
    switch_name = prompt[:-1]
    # switch_name = 'test'

    # Gather data
    script_tab.Screen.Send("term len 0\n")
    script_tab.Screen.WaitForString('\n')
    mac_output = CaptureOutputOfCommand('show mac add dyn', prompt)
    arp_output = CaptureOutputOfCommand('show arp | ex Incomplete', prompt)
    script_tab.Screen.Send("term no len\n")

    # Prep the output file
    filename = switch_name + '.xlsx'
    filename = os.path.join(os.environ['TMPDIR'], filename)
    wb = Workbook()
    wb.save(filename=filename)
    ws1 = wb.active

    # Save MAC table output
    ws1.title = 'MAC_Table'
    ws1['A1'] = 'Vendor'
    ws1['B1'] = 'MAC Address'
    ws1['C1'] = 'Vlan'
    ws1['D1'] = 'Interface'
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 13
    ws1.column_dimensions['E'].width = 18
    ws1['E1'] = datetime.datetime.now()
    mac_index = 1
    for row in mac_output.splitlines():
        if 'DYNAMIC' in row:
            mac_index += 1
            row = row.split()
            vlan = row[0]
            mac_addr = row[1]
            mac_interface = row[3]
            ws1['A' + str(mac_index)] =\
                '=VLOOKUP(LEFT(B%s,7),\
\'/Users/bklotz/Documents/OUI_Table.xlsx\'!Vendor_Table,2,FALSE)' \
% mac_index
            ws1['B' + str(mac_index)] = mac_addr
            ws1['C' + str(mac_index)] = vlan
            ws1['D' + str(mac_index)] = mac_interface

    # Save ARP output
    ws2 = wb.create_sheet(title='ARP_Table')
    ws2['A1'] = 'Vendor'
    ws2['B1'] = 'Address'
    ws2['C1'] = 'Hardware Addr'
    ws2['D1'] = 'Interface'
    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 13
    ws2.column_dimensions['C'].width = 13
    ws2.column_dimensions['E'].width = 18
    ws2['E1'] = datetime.datetime.now()

    arp_index = 1  # Start index at 1 so that data starts in row 2
    for row in arp_output.splitlines():
        if 'Internet' in row:
            arp_index += 1
            row = row.split()
            ip_addr = row[1]
            arp_mac = row[3]
            interface = row[5]
            ws2['A' + str(arp_index)] =\
                '=VLOOKUP(LEFT(C%s,7),\
            \'/Users/bklotz/Documents/OUI_Table.xlsx\'!Vendor_Table,2,FALSE)' \
            % arp_index
            ws2['B' + str(arp_index)] = ip_addr
            ws2['C' + str(arp_index)] = arp_mac
            ws2['D' + str(arp_index)] = interface

    # Add conditional formatting to highlight MACs not in ARP table
    red_text = Font(color="660000")
    red_fill = PatternFill(bgColor="FFCCCC")
    frmla = 'ISNA(VLOOKUP(B2,ARP_Table!$C:$C,1,0))'
    ws1.conditional_formatting.add('$B2:$B' + str(mac_index),
                                   FormulaRule(formula=[frmla],
                                               font=red_text, fill=red_fill))
    # Save spreadsheet
    wb.save(filename)
    # Open spreadsheet
    os.system('open %s' % filename)


def CaptureOutputOfCommand(command, prompt):
    output = ''
    script_tab.Screen.Send(command + '\n')
    script_tab.Screen.WaitForString('\n')
    output = script_tab.Screen.ReadString(prompt)
    #script_tab.Screen.Synchronous = False
    return output


main()
